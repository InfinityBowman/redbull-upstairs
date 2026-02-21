#!/usr/bin/env python3
"""
clean_data.py — Process raw datasets into frontend-ready JSON/GeoJSON.

Reads from python/data/raw/ (populated by fetch_raw.py) and writes
processed files to public/data/.

Usage:
  cd python/
  uv run python scripts/fetch_raw.py   # download first (if not already done)
  uv run python scripts/clean_data.py  # then process

Set DATA_YEAR env var to change the target year (default: 2025).
"""

import csv
import io
import json
import math
import os
import shutil
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

try:
    import requests
except ImportError:
    requests = None  # only needed for weather API

try:
    import shapefile  # pyshp
except ImportError:
    sys.exit("Missing dependency: uv sync")

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: uv sync")

# ── Config ───────────────────────────────────────────────────────────────────

PYTHON_DIR = Path(__file__).resolve().parent.parent  # python/
ROOT = PYTHON_DIR.parent  # repo root
RAW_DIR = PYTHON_DIR / "data" / "raw"
OUT_DIR = ROOT / "public" / "data"
YEAR = int(os.environ.get("DATA_YEAR", "2025"))

STL_COUNTY_FIPS = "29510"

# ── Helpers ──────────────────────────────────────────────────────────────────

def log(msg: str):
    print(f"  → {msg}")


def require_raw(path: Path, name: str):
    """Exit with a helpful message if raw data is missing."""
    if not path.exists():
        sys.exit(f"Missing raw data: {path}\nRun `uv run python scripts/fetch_raw.py` first.")


def web_mercator_to_lnglat(x: float, y: float) -> tuple[float, float]:
    """Convert Web Mercator (EPSG:3857) coordinates to lon/lat (EPSG:4326)."""
    lng = x * 180.0 / 20037508.34
    lat = math.atan(math.exp(y * math.pi / 20037508.34)) * 360.0 / math.pi - 90.0
    return lng, lat


def shapefile_to_geojson(shp_path: str) -> dict:
    """Convert a shapefile to GeoJSON FeatureCollection."""
    sf = shapefile.Reader(shp_path)
    fields = [f[0] for f in sf.fields[1:]]
    features = []
    for sr in sf.shapeRecords():
        props = dict(zip(fields, sr.record))
        for k, v in props.items():
            if isinstance(v, (bytes, bytearray)):
                props[k] = v.decode("utf-8", errors="replace")
        geom = sr.shape.__geo_interface__
        features.append({"type": "Feature", "properties": props, "geometry": geom})
    return {"type": "FeatureCollection", "features": features}


def safe_int(v, default=0):
    try:
        return int(v) if v is not None and str(v).strip().upper() != "NULL" else default
    except (ValueError, TypeError):
        return default


def safe_float(v, default=0.0):
    try:
        return float(v) if v is not None and str(v).strip().upper() != "NULL" else default
    except (ValueError, TypeError):
        return default


# ── 1. CSB 311 Data ──────────────────────────────────────────────────────────

def process_csb() -> None:
    """Process CSB 311 complaint CSVs from raw data."""
    csb_dir = RAW_DIR / "csb"
    require_raw(csb_dir, "CSB")

    csv_files = list(csb_dir.rglob("*.csv"))
    if not csv_files:
        sys.exit("No CSV files found in raw/csb/")

    log(f"Found {len(csv_files)} CSV file(s), parsing...")

    all_rows = []
    for cf in csv_files:
        with open(cf, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            for row in reader:
                all_rows.append(row)

    log(f"Total rows: {len(all_rows):,}")

    # Identify columns
    sample = all_rows[0] if all_rows else {}
    date_col = next((k for k in sample if k.upper() == "DATETIMEINIT"), None)
    if not date_col:
        date_col = next((k for k in sample if "date" in k.lower() and "request" in k.lower()), None)
    if not date_col:
        date_col = next((k for k in sample if "date" in k.lower() and "init" in k.lower()), None)
    if not date_col:
        date_col = next((k for k in sample if "date" in k.lower()), None)
    cat_col = next((k for k in sample if k.upper() == "PROBLEMCODE"), None)
    if not cat_col:
        cat_col = next((k for k in sample if "problem" in k.lower() or "category" in k.lower()), None)
    if not cat_col:
        cat_col = next((k for k in sample if "type" in k.lower()), None)
    status_col = next((k for k in sample if "status" in k.lower()), None)
    hood_col = next((k for k in sample if "neighborhood" in k.lower() or "nhd" in k.lower()), None)
    lat_col = next((k for k in sample if k.lower() in ("latitude", "lat", "y")), None)
    lng_col = next((k for k in sample if k.lower() in ("longitude", "lng", "lon", "long", "x")), None)
    srx_col = next((k for k in sample if k.upper() == "SRX"), None)
    sry_col = next((k for k in sample if k.upper() == "SRY"), None)
    close_date_col = next((k for k in sample if k.upper() == "DATETIMECLOSED"), None)
    if not close_date_col:
        close_date_col = next((k for k in sample if "close" in k.lower() and "date" in k.lower()), None)

    coords_src = "lat/lon" if (lat_col and lng_col) else ("SRX/SRY" if (srx_col and sry_col) else "none")
    log(f"Columns: date={date_col}, category={cat_col}, status={status_col}, hood={hood_col}, coords={coords_src}")

    def parse_date(s: str) -> str | None:
        if not s:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None

    def parse_datetime(s: str) -> datetime | None:
        if not s:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s.strip(), fmt)
            except ValueError:
                continue
        return None

    # Filter to target year
    year_rows = []
    for row in all_rows:
        d = parse_date(row.get(date_col, "")) if date_col else None
        if d and d.startswith(str(YEAR)):
            year_rows.append(row)

    log(f"Rows for {YEAR}: {len(year_rows):,}")

    if not year_rows:
        log(f"WARNING: No rows found for year {YEAR}. Using all data instead.")
        year_rows = all_rows

    # Aggregations
    categories = Counter()
    daily_counts = Counter()
    hourly = Counter()
    weekday = Counter()
    monthly = defaultdict(Counter)
    neighborhoods = defaultdict(lambda: {
        "name": "", "total": 0, "closed": 0, "avgResolutionDays": 0,
        "topCategories": Counter(), "_resolution_days": [],
    })
    heatmap_points = []

    for row in year_rows:
        cat = (row.get(cat_col, "") if cat_col else "").strip() or "Unknown"
        categories[cat] += 1

        date_str = parse_date(row.get(date_col, "")) if date_col else None
        if date_str:
            daily_counts[date_str] += 1
            month_key = date_str[:7]
            monthly[month_key][cat] += 1

        dt = parse_datetime(row.get(date_col, "")) if date_col else None
        if dt:
            hourly[str(dt.hour)] += 1
            weekday[str(dt.weekday())] += 1

        # Neighborhood
        hood_name = (row.get(hood_col, "") if hood_col else "").strip()
        if hood_name:
            nb = neighborhoods[hood_name]
            nb["name"] = hood_name
            nb["total"] += 1
            nb["topCategories"][cat] += 1

            status = (row.get(status_col, "") if status_col else "").strip().lower()
            if "closed" in status or "complete" in status:
                nb["closed"] += 1

            if close_date_col and date_col:
                open_dt = parse_datetime(row.get(date_col, ""))
                close_dt = parse_datetime(row.get(close_date_col, ""))
                if open_dt and close_dt and close_dt > open_dt:
                    days = (close_dt - open_dt).days
                    if days < 365:
                        nb["_resolution_days"].append(days)

        # Heatmap points
        lat, lng = None, None
        if lat_col and lng_col:
            try:
                lat = float(row.get(lat_col, ""))
                lng = float(row.get(lng_col, ""))
            except (ValueError, TypeError):
                pass
        if lat is None and srx_col and sry_col:
            try:
                sx = float(row.get(srx_col, ""))
                sy = float(row.get(sry_col, ""))
                if sx != 0 and sy != 0:
                    lng, lat = web_mercator_to_lnglat(sx, sy)
            except (ValueError, TypeError):
                pass
        if lat is not None and lng is not None:
            if 38.0 < lat < 39.0 and -91.0 < lng < -89.0:
                heatmap_points.append([lat, lng, cat])

    # Finalize neighborhoods
    final_hoods = {}
    for i, (key, nb) in enumerate(sorted(neighborhoods.items(), key=lambda x: -x[1]["total"])):
        hood_id = str(i + 1).zfill(2)
        res_days = nb.pop("_resolution_days", [])
        avg_res = round(sum(res_days) / len(res_days), 1) if res_days else 0
        top_cats = dict(nb["topCategories"].most_common(5))
        final_hoods[hood_id] = {
            "name": nb["name"],
            "total": nb["total"],
            "closed": nb["closed"],
            "avgResolutionDays": avg_res,
            "topCategories": top_cats,
        }

    monthly_out = {}
    for month_key, cats in sorted(monthly.items()):
        monthly_out[month_key] = dict(cats.most_common(10))

    csb_data = {
        "year": YEAR,
        "totalRequests": sum(categories.values()),
        "categories": dict(categories.most_common()),
        "neighborhoods": final_hoods,
        "dailyCounts": dict(sorted(daily_counts.items())),
        "hourly": dict(sorted(hourly.items(), key=lambda x: int(x[0]))),
        "weekday": dict(sorted(weekday.items(), key=lambda x: int(x[0]))),
        "heatmapPoints": heatmap_points[:50000],
        "monthly": monthly_out,
    }

    out_path = OUT_DIR / f"csb_{YEAR}.json"
    with open(out_path, "w") as f:
        json.dump(csb_data, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({out_path.stat().st_size // 1024}KB)")

    latest_path = OUT_DIR / "csb_latest.json"
    shutil.copy2(out_path, latest_path)
    log(f"Copied to {latest_path.name}")

    # ── trends.json (multi-year) ──
    yearly_monthly = defaultdict(dict)
    yearly_categories = defaultdict(Counter)

    for row in all_rows:
        d = parse_date(row.get(date_col, "")) if date_col else None
        if not d:
            continue
        year = d[:4]
        if year not in (str(YEAR), str(YEAR - 1), str(YEAR - 2)):
            continue
        month_key = d[:7]
        cat = (row.get(cat_col, "") if cat_col else "").strip() or "Unknown"
        yearly_monthly[year][month_key] = yearly_monthly[year].get(month_key, 0) + 1
        yearly_categories[year][cat] += 1

    weather_data = fetch_weather(YEAR)

    trends = {
        "yearlyMonthly": {y: dict(sorted(m.items())) for y, m in sorted(yearly_monthly.items())},
        "yearlyCategories": {y: dict(c.most_common()) for y, c in sorted(yearly_categories.items())},
        "weather": weather_data,
    }

    out_path = OUT_DIR / "trends.json"
    with open(out_path, "w") as f:
        json.dump(trends, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({out_path.stat().st_size // 1024}KB)")


def fetch_weather(year: int) -> dict:
    """Fetch daily weather from Open-Meteo historical API for St. Louis."""
    if requests is None:
        log("WARNING: requests not available, skipping weather")
        return {}
    log("Fetching weather data from Open-Meteo...")
    url = (
        f"https://archive-api.open-meteo.com/v1/archive?"
        f"latitude=38.627&longitude=-90.199"
        f"&start_date={year}-01-01&end_date={year}-12-31"
        f"&daily=temperature_2m_max,temperature_2m_min,precipitation_sum"
        f"&temperature_unit=fahrenheit&precipitation_unit=inch"
        f"&timezone=America/Chicago"
    )
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        daily = data.get("daily", {})
        dates = daily.get("time", [])
        highs = daily.get("temperature_2m_max", [])
        lows = daily.get("temperature_2m_min", [])
        precip = daily.get("precipitation_sum", [])

        weather = {}
        for i, d in enumerate(dates):
            weather[d] = {
                "high": round(highs[i], 1) if highs[i] is not None else None,
                "low": round(lows[i], 1) if lows[i] is not None else None,
                "precip": round(precip[i], 2) if precip[i] is not None else 0,
            }
        log(f"Got {len(weather)} days of weather data")
        return weather
    except Exception as e:
        log(f"WARNING: Could not fetch weather data: {e}")
        return {}


# ── 2. Neighborhood Boundaries ───────────────────────────────────────────────

def process_neighborhoods() -> None:
    """Convert neighborhood shapefiles to GeoJSON (reprojected to WGS84)."""
    import geopandas as gpd

    nhd_dir = RAW_DIR / "neighborhoods"
    require_raw(nhd_dir, "neighborhoods")

    shp_files = list(nhd_dir.rglob("*.shp"))
    if not shp_files:
        sys.exit("No .shp files found in raw/neighborhoods/")

    log(f"Converting {shp_files[0].name} to GeoJSON...")
    gdf = gpd.read_file(shp_files[0])

    if gdf.crs and gdf.crs != "EPSG:4326":
        log(f"Reprojecting from {gdf.crs} to EPSG:4326...")
        gdf = gdf.to_crs(epsg=4326)

    geojson = json.loads(gdf.to_json())
    log(f"{len(geojson['features'])} neighborhood features")

    out_path = OUT_DIR / "neighborhoods.geojson"
    with open(out_path, "w") as f:
        json.dump(geojson, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({out_path.stat().st_size // 1024}KB)")


# ── 3. Transit (GTFS) ───────────────────────────────────────────────────────

def process_gtfs() -> None:
    """Process GTFS feed into stops, routes, shapes, stop_stats."""
    gtfs_dir = RAW_DIR / "gtfs"
    require_raw(gtfs_dir, "GTFS")

    # Also check for the zip file directly
    gtfs_zip = RAW_DIR / "google_transit.zip"
    if gtfs_zip.exists():
        zf = zipfile.ZipFile(gtfs_zip)
    else:
        # Look for txt files directly in the extracted dir
        zf = None

    def open_gtfs_file(name: str):
        if zf and name in zf.namelist():
            return io.TextIOWrapper(zf.open(name), encoding="utf-8-sig")
        path = gtfs_dir / name
        if path.exists():
            return open(path, "r", encoding="utf-8-sig")
        return None

    def has_gtfs_file(name: str) -> bool:
        if zf and name in zf.namelist():
            return True
        return (gtfs_dir / name).exists()

    log("Parsing GTFS feed...")

    # ── stops.geojson ──
    f = open_gtfs_file("stops.txt")
    if not f:
        sys.exit("No stops.txt found in GTFS data")
    reader = csv.DictReader(f)
    stops = list(reader)
    f.close()

    features = []
    for s in stops:
        try:
            lat = float(s.get("stop_lat", 0))
            lon = float(s.get("stop_lon", 0))
        except ValueError:
            continue
        if lat == 0 or lon == 0:
            continue
        features.append({
            "type": "Feature",
            "properties": {
                "stop_id": s.get("stop_id", ""),
                "stop_name": s.get("stop_name", ""),
                "stop_code": s.get("stop_code", ""),
            },
            "geometry": {"type": "Point", "coordinates": [lon, lat]},
        })

    stops_geo = {"type": "FeatureCollection", "features": features}
    out_path = OUT_DIR / "stops.geojson"
    with open(out_path, "w") as f:
        json.dump(stops_geo, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({len(features)} stops, {out_path.stat().st_size // 1024}KB)")

    # ── routes.json ──
    f = open_gtfs_file("routes.txt")
    if f:
        reader = csv.DictReader(f)
        routes_list = []
        for r in reader:
            routes_list.append({
                "route_id": r.get("route_id", ""),
                "route_short_name": r.get("route_short_name", ""),
                "route_long_name": r.get("route_long_name", ""),
                "route_type": int(r.get("route_type", 3)),
                "route_color": f"#{r['route_color']}" if r.get("route_color") else "",
            })
        f.close()

        out_path = OUT_DIR / "routes.json"
        with open(out_path, "w") as f:
            json.dump(routes_list, f, separators=(",", ":"))
        log(f"Wrote {out_path.name} ({len(routes_list)} routes)")

    # ── shapes.geojson ──
    if has_gtfs_file("shapes.txt"):
        f = open_gtfs_file("shapes.txt")
        reader = csv.DictReader(f)
        shape_points = defaultdict(list)
        for row in reader:
            sid = row.get("shape_id", "")
            try:
                lat = float(row.get("shape_pt_lat", 0))
                lon = float(row.get("shape_pt_lon", 0))
                seq = int(row.get("shape_pt_sequence", 0))
            except ValueError:
                continue
            shape_points[sid].append((seq, [lon, lat]))
        f.close()

        shape_to_route = {}
        if has_gtfs_file("trips.txt"):
            f = open_gtfs_file("trips.txt")
            reader = csv.DictReader(f)
            for row in reader:
                sid = row.get("shape_id", "")
                rid = row.get("route_id", "")
                if sid and rid:
                    shape_to_route[sid] = rid
            f.close()

        shape_features = []
        for sid, pts in shape_points.items():
            pts.sort(key=lambda x: x[0])
            coords = [p[1] for p in pts]
            if len(coords) < 2:
                continue
            shape_features.append({
                "type": "Feature",
                "properties": {"shape_id": sid, "route_id": shape_to_route.get(sid, "")},
                "geometry": {"type": "LineString", "coordinates": coords},
            })

        shapes_geo = {"type": "FeatureCollection", "features": shape_features}
        out_path = OUT_DIR / "shapes.geojson"
        with open(out_path, "w") as f:
            json.dump(shapes_geo, f, separators=(",", ":"))
        log(f"Wrote {out_path.name} ({len(shape_features)} shapes, {out_path.stat().st_size // 1024}KB)")

    # ── stop_stats.json ──
    if has_gtfs_file("stop_times.txt"):
        trip_to_route = {}
        if has_gtfs_file("trips.txt"):
            f = open_gtfs_file("trips.txt")
            reader = csv.DictReader(f)
            for row in reader:
                trip_to_route[row.get("trip_id", "")] = row.get("route_id", "")
            f.close()

        stop_trips = defaultdict(set)
        stop_routes = defaultdict(set)

        f = open_gtfs_file("stop_times.txt")
        reader = csv.DictReader(f)
        for row in reader:
            sid = row.get("stop_id", "")
            tid = row.get("trip_id", "")
            stop_trips[sid].add(tid)
            rid = trip_to_route.get(tid, "")
            if rid:
                stop_routes[sid].add(rid)
        f.close()

        stats = {}
        for sid in stop_trips:
            stats[sid] = {
                "trip_count": len(stop_trips[sid]),
                "routes": sorted(stop_routes.get(sid, set())),
            }

        out_path = OUT_DIR / "stop_stats.json"
        with open(out_path, "w") as f:
            json.dump(stats, f, separators=(",", ":"))
        log(f"Wrote {out_path.name} ({len(stats)} stops, {out_path.stat().st_size // 1024}KB)")

    if zf:
        zf.close()


# ── 4. Food Desert Tracts ────────────────────────────────────────────────────

def process_food_deserts() -> None:
    """Merge USDA food access data with Census TIGER tract geometries."""
    xlsx_path = RAW_DIR / "food-access-research-atlas-data-download-2019.xlsx"
    require_raw(xlsx_path, "USDA food access")

    tiger_dir = RAW_DIR / "tiger_tracts"
    require_raw(tiger_dir, "TIGER tracts")

    log("Parsing USDA Food Access Research Atlas...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Food Access Research Atlas"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    tract_col = next((i for i, h in enumerate(headers) if h and "CensusTract" in str(h)), None)
    pop_col = next((i for i, h in enumerate(headers) if h and str(h).strip() == "POP2010"), None)
    poverty_col = next((i for i, h in enumerate(headers) if h and "Poverty" in str(h) and "Rate" in str(h)), None)
    lila_col = next((i for i, h in enumerate(headers) if h and str(h).strip() in ("LILATracts_1And10", "LILA1and10")), None)
    vehicle_col = next((i for i, h in enumerate(headers) if h and "lahunv" in str(h).lower()), None)
    income_col = next((i for i, h in enumerate(headers) if h and "Median" in str(h) and "Income" in str(h)), None)

    stl_tracts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tract_id = str(row[tract_col]) if tract_col is not None and row[tract_col] else ""
        if not tract_id.startswith(STL_COUNTY_FIPS):
            continue

        pop = safe_float(row[pop_col] if pop_col is not None else 0)
        poverty = safe_float(row[poverty_col] if poverty_col is not None else 0)
        lila = bool(row[lila_col]) if lila_col is not None and row[lila_col] is not None else False
        vehicle_pct = 0
        if vehicle_col is not None and row[vehicle_col] and pop > 0:
            vehicle_pct = round(safe_float(row[vehicle_col]) / pop * 100, 1)
        median_income = safe_int(row[income_col] if income_col is not None else 0)

        stl_tracts[tract_id] = {
            "pop": safe_int(pop),
            "poverty_rate": round(poverty, 1),
            "lila": lila,
            "pct_no_vehicle": vehicle_pct,
            "median_income": median_income,
        }

    wb.close()
    log(f"Found {len(stl_tracts)} St. Louis census tracts in USDA data")

    shp_files = list(tiger_dir.rglob("*.shp"))
    if not shp_files:
        sys.exit("No .shp files in raw/tiger_tracts/")

    log("Reading TIGER/Line tract geometries...")
    sf = shapefile.Reader(str(shp_files[0]))
    fields = [f[0] for f in sf.fields[1:]]
    geoid_idx = fields.index("GEOID") if "GEOID" in fields else 0
    name_idx = fields.index("NAMELSAD") if "NAMELSAD" in fields else 1

    features = []
    for sr in sf.shapeRecords():
        geoid = str(sr.record[geoid_idx])
        if not geoid.startswith(STL_COUNTY_FIPS):
            continue

        tract_data = stl_tracts.get(geoid, {})
        name = sr.record[name_idx]

        features.append({
            "type": "Feature",
            "properties": {
                "tract_id": geoid,
                "name": name,
                "poverty_rate": tract_data.get("poverty_rate", 0),
                "pop": tract_data.get("pop", 0),
                "pct_no_vehicle": tract_data.get("pct_no_vehicle", 0),
                "nearest_grocery_miles": 1.5,  # placeholder — app computes at runtime
                "lila": tract_data.get("lila", False),
                "median_income": tract_data.get("median_income", 0),
            },
            "geometry": sr.shape.__geo_interface__,
        })

    food_geo = {"type": "FeatureCollection", "features": features}
    out_path = OUT_DIR / "food_deserts.geojson"
    with open(out_path, "w") as f:
        json.dump(food_geo, f, separators=(",", ":"))
    log(f"Wrote {out_path.name} ({len(features)} tracts, {out_path.stat().st_size // 1024}KB)")


# ── 5. Grocery Stores (embedded) ─────────────────────────────────────────────

GROCERY_STORES = [
    {"name": "Schnucks Arsenal", "chain": "Schnucks", "coords": [-90.2418, 38.5937]},
    {"name": "Schnucks Loughborough", "chain": "Schnucks", "coords": [-90.2689, 38.5830]},
    {"name": "Schnucks Hampton Village", "chain": "Schnucks", "coords": [-90.2920, 38.5917]},
    {"name": "Schnucks Lindbergh", "chain": "Schnucks", "coords": [-90.3360, 38.5803]},
    {"name": "Schnucks Clayton Rd", "chain": "Schnucks", "coords": [-90.3260, 38.6337]},
    {"name": "Schnucks Kirkwood", "chain": "Schnucks", "coords": [-90.4150, 38.5790]},
    {"name": "Schnucks Des Peres", "chain": "Schnucks", "coords": [-90.3990, 38.5970]},
    {"name": "Schnucks Shrewsbury", "chain": "Schnucks", "coords": [-90.3190, 38.5880]},
    {"name": "Dierbergs Brentwood", "chain": "Dierbergs", "coords": [-90.3490, 38.6193]},
    {"name": "Dierbergs Maplewood", "chain": "Dierbergs", "coords": [-90.3160, 38.6100]},
    {"name": "Aldi Gravois", "chain": "Aldi", "coords": [-90.2630, 38.5850]},
    {"name": "Aldi Hampton", "chain": "Aldi", "coords": [-90.2900, 38.6030]},
    {"name": "Aldi Grand", "chain": "Aldi", "coords": [-90.2180, 38.6290]},
    {"name": "Aldi Natural Bridge", "chain": "Aldi", "coords": [-90.2530, 38.6700]},
    {"name": "Save-A-Lot Gravois", "chain": "Save-A-Lot", "coords": [-90.2470, 38.5960]},
    {"name": "Save-A-Lot North Broadway", "chain": "Save-A-Lot", "coords": [-90.1990, 38.6750]},
    {"name": "Save-A-Lot Chippewa", "chain": "Save-A-Lot", "coords": [-90.2730, 38.5900]},
    {"name": "Ruler Foods Lemay Ferry", "chain": "Ruler Foods", "coords": [-90.2800, 38.5430]},
    {"name": "Fresh Thyme Brentwood", "chain": "Fresh Thyme", "coords": [-90.3480, 38.6170]},
    {"name": "Trader Joe's Brentwood", "chain": "Trader Joe's", "coords": [-90.3510, 38.6180]},
    {"name": "Whole Foods Brentwood", "chain": "Whole Foods", "coords": [-90.3520, 38.6160]},
    {"name": "Tower Grove Farmers Market", "chain": "Farmers Market", "coords": [-90.2520, 38.6100]},
    {"name": "Soulard Farmers Market", "chain": "Farmers Market", "coords": [-90.2080, 38.6130]},
]


def write_grocery_stores() -> None:
    """Write embedded grocery store data as GeoJSON."""
    features = [
        {
            "type": "Feature",
            "properties": {"name": s["name"], "chain": s["chain"]},
            "geometry": {"type": "Point", "coordinates": s["coords"]},
        }
        for s in GROCERY_STORES
    ]
    geo = {"type": "FeatureCollection", "features": features}
    out_path = OUT_DIR / "grocery_stores.geojson"
    with open(out_path, "w") as f:
        json.dump(geo, f, indent=2)
    log(f"Wrote {out_path.name} ({len(features)} stores)")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  STL Urban Analytics — Data Cleaner")
    print(f"  Raw input:  {RAW_DIR}")
    print(f"  Output:     {OUT_DIR}")
    print(f"  Target year: {YEAR}")
    print("=" * 60)

    if not RAW_DIR.exists():
        sys.exit(f"\nNo raw data found at {RAW_DIR}\nRun `uv run python scripts/fetch_raw.py` first.")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    steps = [
        ("Neighborhoods", process_neighborhoods),
        ("GTFS transit", process_gtfs),
        ("Food deserts", process_food_deserts),
        ("Grocery stores", write_grocery_stores),
        ("CSB 311 data", process_csb),
    ]

    for name, fn in steps:
        try:
            print(f"\n── {name} ──")
            fn()
        except SystemExit:
            raise
        except Exception as e:
            print(f"\n❌ {name} failed: {e}")

    # Summary
    print("\n" + "=" * 60)
    print("  Done! Files in public/data/:")
    print("=" * 60)
    for f in sorted(OUT_DIR.iterdir()):
        size = f.stat().st_size
        unit = "KB" if size > 1024 else "B"
        val = size // 1024 if size > 1024 else size
        print(f"  {f.name:<30} {val:>6} {unit}")


if __name__ == "__main__":
    main()
